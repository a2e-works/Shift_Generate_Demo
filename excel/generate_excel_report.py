"""
管理者向けExcelレポートの生成

目的:
  管理者への提示は使い慣れたExcel形式が望ましいという要望に応え、
  「氏名を左列、日付を横に並べたシフト表」+「シフト健全度」を1つのExcelにまとめる。

  重要: シフト健全度はハードコードされた数値ではなく、メンバーデータ(資格・
  設備別スキル段階・バックアップ可否)からExcelの数式で自動計算する。
  これにより「教育を進める・資格を取得する」とスコアが実際に上がることを
  Excel上でそのまま確認でき、管理者だけでなくメンバー自身にも
  「教育すると希望休が通りやすくなる」ことを納得してもらいやすくする狙いがある。

シート構成:
  1. シフト健全度        … 総合スコア+5指標(数式)、判定基準、今月の早番A活用状況
  2. シフト表(2026年8月) … 氏名×日付のシフト表(現場のExcelそのままのイメージ)
  3. 希望休申請一覧      … 各申請について、承認した場合のチーム残人数と判定
  4. 頭打ち箇所と対応案  … 教育担当が不在で頭打ちの箇所+3つの対応案(STEP7と連動)
  5. メンバーデータ      … 元データ+数値化した段階(数式のソース)
  6. 段階マップ(非表示) … 段階名→数値の対応表
"""

from __future__ import annotations

import json
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
from members import load_members as _load_members_objs  # noqa: E402
from education import load_roadmap  # noqa: E402
from shift_schedule import generate_schedule, DEMO_START_DATE, DEMO_NUM_DAYS  # noqa: E402
from auto_approval import evaluate_all_requests, find_weekly_hour_violations  # noqa: E402
from substitution import simulate_substitution  # noqa: E402
from future_simulation import (  # noqa: E402
    instructor_gap_report,
    distinct_promotion_need,
    gap_resolution_options,
)

BASE = Path(__file__).resolve().parent.parent
DATA_DIR = BASE / "data"
OUT_PATH = BASE / "excel" / "A2EWorks_シフト健全度レポート_2026年8月.xlsx"

STAGE_ORDER = ["見学", "補助", "単独", "教育担当補助", "教育担当"]
EQUIPMENT_LIST = ["受変電設備", "空調設備", "消防設備", "監視盤"]
ROTATION_CYCLE = ["早A", "早B", "夜", "明", "休"]
TEAM_OFFSETS = {"A": 0, "B": 1, "C": 2, "D": 3, "E": 4}
START_DATE = date(2026, 8, 1)
NUM_DAYS = 31
WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]

FONT_NAME = "Arial"

SHIFT_FILL = {
    "早A": PatternFill("solid", fgColor="E4D6F7"),  # 教育投資枠(紫系)
    "早B": PatternFill("solid", fgColor="FDECC8"),  # 早番(琥珀系)
    "夜": PatternFill("solid", fgColor="D6E4F0"),   # 夜勤(青系)
    "明": PatternFill("solid", fgColor="E8E8E8"),   # 明け(グレー)
    "休": PatternFill("solid", fgColor="D9F2E3"),   # 休み(緑系)
}
REQUEST_BORDER = Border(*(Side(style="thick", color="C0392B") for _ in range(4)))

HEADER_FILL = PatternFill("solid", fgColor="2F3B4C")
HEADER_FONT = Font(name=FONT_NAME, color="FFFFFF", bold=True)
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True)
BOLD = Font(name=FONT_NAME, bold=True)
NORMAL = Font(name=FONT_NAME)


def load_json(name):
    with open(DATA_DIR / name, "r", encoding="utf-8") as f:
        return json.load(f)


def shift_for_team(day_index: int, team: str) -> str:
    return ROTATION_CYCLE[(day_index + TEAM_OFFSETS[team]) % 5]


def _est_lines(text, col_width):
    """列幅とテキスト量から、折り返し後の行数を概算する(日本語は1文字≒2ユニット幅として計算)。"""
    if not text:
        return 1
    capacity = max(4, col_width / 2.0)
    return max(1, -(-len(text) // int(capacity)))


def _fit_row_height(ws, row, texts_and_widths, base=16, pad=8):
    max_lines = max(_est_lines(t, w) for t, w in texts_and_widths) if texts_and_widths else 1
    ws.row_dimensions[row].height = base * max_lines + pad


def build_workbook():
    members = load_json("members.json")
    constraints = load_json("constraints.json")
    member_objs = _load_members_objs()
    member_objs_by_id = {m.member_id: m for m in member_objs}
    roadmap = load_roadmap()
    schedule_for_auto_approval = generate_schedule(member_objs, DEMO_START_DATE, DEMO_NUM_DAYS)
    auto_approval_results = {
        (r["member_id"], r["date"]): r for r in evaluate_all_requests(member_objs, schedule_for_auto_approval, constraints)
    }

    wb = Workbook()
    wb.remove(wb.active)

    ws_guide = wb.create_sheet("使い方")
    ws_health = wb.create_sheet("シフト健全度")
    ws_shift = wb.create_sheet("シフト表(2026年8月)")
    ws_requests = wb.create_sheet("希望休申請一覧")
    ws_hours = wb.create_sheet("労働時間・自動承認設定")
    ws_gaps = wb.create_sheet("頭打ち箇所と対応案")
    ws_members = wb.create_sheet("メンバーデータ")
    ws_stagemap = wb.create_sheet("段階マップ")

    # ---------------- 段階マップ ----------------
    ws_stagemap["A1"] = "段階名"
    ws_stagemap["B1"] = "数値"
    for i, stage in enumerate(STAGE_ORDER):
        ws_stagemap.cell(row=i + 2, column=1, value=stage)
        ws_stagemap.cell(row=i + 2, column=2, value=i)
    ws_stagemap.sheet_state = "hidden"

    # ---------------- メンバーデータ ----------------
    headers = [
        "氏名", "チーム", "役職", "経験月数", "資格",
        "受変電設備段階", "空調設備段階", "消防設備段階", "監視盤段階",
        "バックアップ可能業務",
        "受変電_数値", "空調_数値", "消防_数値", "監視盤_数値",
        "いずれかで単独以上か",
    ]
    for c, h in enumerate(headers, start=1):
        cell = ws_members.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws_members.freeze_panes = "A2"

    for r, m in enumerate(members, start=2):
        ws_members.cell(row=r, column=1, value=m["name"])
        ws_members.cell(row=r, column=2, value=m["team"])
        ws_members.cell(row=r, column=3, value=m["role"])
        ws_members.cell(row=r, column=4, value=m["experience_months"])
        ws_members.cell(row=r, column=5, value=", ".join(m["certifications"]))
        for ci, eq in enumerate(EQUIPMENT_LIST):
            ws_members.cell(row=r, column=6 + ci, value=m["equipment_skills"].get(eq, ""))
        ws_members.cell(row=r, column=10, value=", ".join(m["backup_capable_tasks"]))
        # 数値化(段階マップをVLOOKUP)
        for ci, col_letter in zip(range(4), ["F", "G", "H", "I"]):
            out_col = 11 + ci
            ws_members.cell(
                row=r, column=out_col,
                value=f"=VLOOKUP({col_letter}{r},段階マップ!$A:$B,2,FALSE)"
            )
        ws_members.cell(row=r, column=15, value=f"=IF(MAX(K{r}:N{r})>=2,1,0)")

    last_member_row = len(members) + 1  # 16

    for col_idx in range(1, 16):
        ws_members.column_dimensions[get_column_letter(col_idx)].width = 14

    # ---------------- シフト表(2026年8月) ----------------
    ws_shift["A1"] = "氏名"
    ws_shift["B1"] = "チーム"
    ws_shift["C1"] = "役職"
    for c in ("A1", "B1", "C1"):
        ws_shift[c].font = HEADER_FONT
        ws_shift[c].fill = HEADER_FILL

    date_col_start = 4
    for i in range(NUM_DAYS):
        d = START_DATE + timedelta(days=i)
        col = date_col_start + i
        cell = ws_shift.cell(row=1, column=col, value=d)
        cell.number_format = 'm/d("' + WEEKDAY_JP[d.weekday()] + '")'
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws_shift.column_dimensions[get_column_letter(col)].width = 6

    count_col = date_col_start + NUM_DAYS  # 早A回数列
    hdr = ws_shift.cell(row=1, column=count_col, value="早A回数(今月)")
    hdr.font = HEADER_FONT
    hdr.fill = HEADER_FILL
    ws_shift.column_dimensions[get_column_letter(count_col)].width = 14

    requested_days_off = constraints.get("requested_days_off", {})
    member_id_to_name = {}
    members_by_name = {}
    for m in members:
        member_id_to_name[m["member_id"]] = m["name"]
        members_by_name[m["name"]] = m

    requested_dates_by_name = {}
    for mid, dates in requested_days_off.items():
        requested_dates_by_name[member_id_to_name[mid]] = set(dates)

    for r, m in enumerate(members, start=2):
        ws_shift.cell(row=r, column=1, value=m["name"])
        ws_shift.cell(row=r, column=2, value=m["team"])
        ws_shift.cell(row=r, column=3, value=m["role"])
        req_dates = requested_dates_by_name.get(m["name"], set())
        for i in range(NUM_DAYS):
            d = START_DATE + timedelta(days=i)
            code = shift_for_team(i, m["team"])
            cell = ws_shift.cell(row=r, column=date_col_start + i, value=code)
            cell.fill = SHIFT_FILL[code]
            cell.font = NORMAL
            cell.alignment = Alignment(horizontal="center")
            if d.isoformat() in req_dates:
                cell.border = REQUEST_BORDER
                cell.comment = Comment("希望休の申請あり", "A2E Works")
        first_col_letter = get_column_letter(date_col_start)
        last_col_letter = get_column_letter(date_col_start + NUM_DAYS - 1)
        ws_shift.cell(
            row=r, column=count_col,
            value=f'=COUNTIF({first_col_letter}{r}:{last_col_letter}{r},"早A")'
        )

    ws_shift.freeze_panes = get_column_letter(date_col_start) + "2"

    legend_row = last_member_row + 2
    ws_shift.cell(row=legend_row, column=1, value="凡例:").font = BOLD
    legend_items = [
        ("早A", "早番A(教育投資枠)"), ("早B", "早番B"), ("夜", "夜勤"),
        ("明", "明け"), ("休", "休み"),
    ]
    for i, (code, label) in enumerate(legend_items):
        cell = ws_shift.cell(row=legend_row + 1 + i, column=1, value=code)
        cell.fill = SHIFT_FILL[code]
        cell.alignment = Alignment(horizontal="center")
        ws_shift.cell(row=legend_row + 1 + i, column=2, value=label).font = NORMAL
    ws_shift.cell(row=legend_row + 1 + len(legend_items), column=1, value="太赤枠").font = BOLD
    ws_shift.cell(row=legend_row + 1 + len(legend_items), column=2, value="希望休の申請あり").font = NORMAL

    # ---------------- 希望休申請一覧 ----------------
    ws_requests["A1"] = "現在のシフト健全度:"
    ws_requests["A1"].font = BOLD
    ws_requests["B1"] = "=シフト健全度!C3"
    ws_requests["B1"].font = BOLD
    ws_requests["B1"].number_format = "0.0"
    ws_requests["C1"] = "=シフト健全度!C4"

    req_headers = ["氏名", "チーム", "希望日", "理由", "割当中のシフト", "承認後のチーム残(単独対応以上)人数", "判定", "申請日", "自動承認可否", "自動承認/手動確認の理由", "必要人数", "交代要員の提案"]
    header_row = 3
    for c, h in enumerate(req_headers, start=1):
        cell = ws_requests.cell(row=header_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws_requests.column_dimensions["A"].width = 14
    ws_requests.column_dimensions["D"].width = 16
    ws_requests.column_dimensions["F"].width = 30
    ws_requests.column_dimensions["G"].width = 26
    ws_requests.column_dimensions["H"].width = 14
    ws_requests.column_dimensions["I"].width = 14
    ws_requests.column_dimensions["J"].width = 55
    ws_requests.column_dimensions["K"].width = 12
    ws_requests.column_dimensions["L"].width = 45

    shift_first_col_letter = get_column_letter(date_col_start)
    shift_last_col_letter = get_column_letter(date_col_start + NUM_DAYS - 1)

    row = header_row + 1
    for mid, dates in requested_days_off.items():
        name = member_id_to_name[mid]
        for d_str in dates:
            ws_requests.cell(row=row, column=1, value=name)
            ws_requests.cell(row=row, column=2, value=f"=VLOOKUP(A{row},メンバーデータ!$A:$B,2,FALSE)")
            date_cell = ws_requests.cell(row=row, column=3, value=date.fromisoformat(d_str))
            d_obj = date.fromisoformat(d_str)
            date_cell.number_format = 'm/d("' + WEEKDAY_JP[d_obj.weekday()] + '")'
            ws_requests.cell(row=row, column=4, value="(初期データ)")
            ws_requests.cell(
                row=row, column=5,
                value=(
                    f"=INDEX('シフト表(2026年8月)'!${shift_first_col_letter}$2:${shift_last_col_letter}$16,"
                    f"MATCH($A{row},'シフト表(2026年8月)'!$A$2:$A$16,0),"
                    f"MATCH($C{row},'シフト表(2026年8月)'!${shift_first_col_letter}$1:${shift_last_col_letter}$1,0))"
                )
            )
            ws_requests.cell(
                row=row, column=6,
                value=f"=SUMIFS(メンバーデータ!$O:$O,メンバーデータ!$B:$B,B{row})-VLOOKUP(A{row},メンバーデータ!$A:$O,15,FALSE)"
            )
            ws_requests.cell(
                row=row, column=7,
                value=f'=IF(F{row}>=2,"バックアップ不要",IF(F{row}=1,"要注意(イベント時はバックアップ検討)","バックアップ出動が必要"))'
            )
            auto_result = auto_approval_results.get((mid, d_str))
            if auto_result:
                submitted_cell = ws_requests.cell(row=row, column=8, value=date.fromisoformat(auto_result["submitted_on"]))
                submitted_cell.number_format = "yyyy/mm/dd"
                verdict_cell = ws_requests.cell(
                    row=row, column=9, value="自動承認" if auto_result["auto_approved"] else "手動確認"
                )
                verdict_cell.font = Font(
                    name=FONT_NAME, bold=True,
                    color="2E7D46" if auto_result["auto_approved"] else "B04A2E",
                )
                reason_cell = ws_requests.cell(row=row, column=10, value=" / ".join(auto_result["reasons"]))
                reason_cell.alignment = Alignment(wrap_text=True, vertical="top")

            sub = simulate_substitution(d_str, member_objs_by_id[mid].team, mid, member_objs, schedule_for_auto_approval, constraints)
            ws_requests.cell(row=row, column=11, value=sub["required_headcount"])
            swap_text = (
                f"{sub['swap_out']} → {sub['swap_in']}(チーム{sub['swap_in_team']})"
                if sub.get("needs_substitute") and sub.get("swap_in")
                else ("交代不要" if not sub.get("needs_substitute") else "交代候補なし(要検討)")
            )
            swap_cell = ws_requests.cell(row=row, column=12, value=swap_text)
            swap_cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    note_row = row + 1
    ws_requests.cell(row=note_row, column=1,
                      value="※「承認後のチーム残人数」は、そのチームで設備を単独対応以上でこなせる人数から、申請者本人の分を除いた人数です。").font = Font(name=FONT_NAME, italic=True, size=9)
    ws_requests.cell(row=note_row + 1, column=1,
                      value="※「自動承認可否」はsrc/auto_approval.pyで計算した値です(申請期限・週40時間・残人数・NGペアをすべて満たす場合のみ自動承認)。設定は「労働時間・自動承認設定」シート参照。").font = Font(name=FONT_NAME, italic=True, size=9)

    # ---------------- シフト健全度 ----------------
    ws_health.column_dimensions["A"].width = 26
    ws_health.column_dimensions["B"].width = 16
    ws_health.column_dimensions["C"].width = 16
    ws_health.column_dimensions["D"].width = 16

    ws_health["A1"] = "シフト健全度レポート(2026年8月時点)"
    ws_health["A1"].font = TITLE_FONT
    ws_health.merge_cells("A1:D1")

    metric_cells = {
        "教育達成率": None,
        "属人化耐性": None,
        "資格充足率": None,
        "バックアップ率": None,
        "変更耐性": None,
    }

    # 各セクションの開始行を固定値で予約し、重なりが絶対に起きないようにする。
    EQ_LABEL_ROW = 15
    EQ_HEADER_ROW = 16
    EQ_DATA_START = 17          # 17-20 (4設備)
    DEPENDENCY_OVERALL_ROW = 21

    TEAM_LABEL_ROW = 23
    TEAM_HEADER_ROW = 24
    TEAM_DATA_START = 25        # 25-29 (5チーム)
    RESILIENCE_OVERALL_ROW = 30

    QUAL_LABEL_ROW = 32
    QUAL_HEADER_ROW = 33
    QUAL_DATA_START = 34        # 34-36 (3設備)
    QUAL_OVERALL_ROW = 37

    SIMPLE_ROW = 39             # 教育達成率
    SIMPLE_ROW2 = 40            # バックアップ率

    # --- 内訳: 属人化耐性(設備別) ---
    ws_health.cell(row=EQ_LABEL_ROW, column=1, value="属人化耐性 内訳(設備別: 単独対応以上の人数)").font = BOLD
    for c, h in enumerate(["設備", "人数", "スコア"], start=1):
        cell = ws_health.cell(row=EQ_HEADER_ROW, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    eq_col_map = {"受変電設備": "K", "空調設備": "L", "消防設備": "M", "監視盤": "N"}
    eq_score_cells = []
    for i, eq in enumerate(EQUIPMENT_LIST):
        r = EQ_DATA_START + i
        col_letter = eq_col_map[eq]
        ws_health.cell(row=r, column=1, value=eq)
        count_formula = f'=COUNTIF(メンバーデータ!{col_letter}2:{col_letter}{last_member_row},">=2")'
        ws_health.cell(row=r, column=2, value=count_formula)
        score_formula = f"=MIN(B{r},3)/3*100"
        score_cell = ws_health.cell(row=r, column=3, value=score_formula)
        score_cell.number_format = "0.0"
        eq_score_cells.append(f"C{r}")

    ws_health.cell(row=DEPENDENCY_OVERALL_ROW, column=1, value="属人化耐性(総合)").font = BOLD
    dep_overall_cell = ws_health.cell(row=DEPENDENCY_OVERALL_ROW, column=3, value=f"=AVERAGE({','.join(eq_score_cells)})")
    dep_overall_cell.number_format = "0.0"
    metric_cells["属人化耐性"] = f"C{DEPENDENCY_OVERALL_ROW}"

    # --- 内訳: 変更耐性(チーム別) ---
    ws_health.cell(row=TEAM_LABEL_ROW, column=1, value="変更耐性 内訳(チーム別: 単独対応可能人数)").font = BOLD
    for c, h in enumerate(["チーム", "単独対応可能人数", "チーム人数", "スコア"], start=1):
        cell = ws_health.cell(row=TEAM_HEADER_ROW, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    team_score_cells = []
    for i, team in enumerate(["A", "B", "C", "D", "E"]):
        r = TEAM_DATA_START + i
        ws_health.cell(row=r, column=1, value=team)
        ws_health.cell(row=r, column=2, value=f'=SUMIFS(メンバーデータ!$O:$O,メンバーデータ!$B:$B,"{team}")')
        ws_health.cell(row=r, column=3, value=f'=COUNTIF(メンバーデータ!$B:$B,"{team}")')
        score_cell2 = ws_health.cell(row=r, column=4, value=f"=MIN(B{r},2)/2*100")
        score_cell2.number_format = "0.0"
        team_score_cells.append(f"D{r}")

    ws_health.cell(row=RESILIENCE_OVERALL_ROW, column=1, value="変更耐性(総合)").font = BOLD
    res_overall_cell = ws_health.cell(row=RESILIENCE_OVERALL_ROW, column=4, value=f"=AVERAGE({','.join(team_score_cells)})")
    res_overall_cell.number_format = "0.0"
    metric_cells["変更耐性"] = f"D{RESILIENCE_OVERALL_ROW}"

    # --- 資格充足率(計算用ワーク行) ---
    ws_health.cell(row=QUAL_LABEL_ROW, column=1, value="資格充足率 内訳(単独対応以上かつ必要資格を保有)").font = BOLD
    for c, h in enumerate(["設備", "対象人数", "資格保有人数"], start=1):
        cell = ws_health.cell(row=QUAL_HEADER_ROW, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    qual_defs = [
        ("受変電設備", "K", ["第二種電気工事士"]),
        ("空調設備", "L", ["第一種冷凍機械責任者", "第三種冷凍機械責任者"]),
        ("消防設備", "M", ["危険物取扱者乙4"]),
    ]
    total_check_cells = []
    satisfied_cells = []
    for i, (eq, col_letter, certs) in enumerate(qual_defs):
        r = QUAL_DATA_START + i
        ws_health.cell(row=r, column=1, value=eq)
        total_formula = f'=COUNTIF(メンバーデータ!{col_letter}2:{col_letter}{last_member_row},">=2")'
        ws_health.cell(row=r, column=2, value=total_formula)
        cert_terms = "+".join(
            f'ISNUMBER(SEARCH("{c}",メンバーデータ!$E$2:$E${last_member_row}))' for c in certs
        )
        satisfied_formula = (
            f"=SUMPRODUCT((メンバーデータ!{col_letter}$2:{col_letter}${last_member_row}>=2)*"
            f"(({cert_terms})>0))"
        )
        ws_health.cell(row=r, column=3, value=satisfied_formula)
        total_check_cells.append(f"B{r}")
        satisfied_cells.append(f"C{r}")

    ws_health.cell(row=QUAL_OVERALL_ROW, column=1, value="資格充足率(総合)").font = BOLD
    ws_health.cell(row=QUAL_OVERALL_ROW, column=2, value=f"=SUM({','.join(total_check_cells)})")
    ws_health.cell(row=QUAL_OVERALL_ROW, column=3, value=f"=SUM({','.join(satisfied_cells)})")
    ws_health.cell(row=QUAL_OVERALL_ROW, column=4, value=f"=C{QUAL_OVERALL_ROW}/B{QUAL_OVERALL_ROW}*100").number_format = "0.0"
    metric_cells["資格充足率"] = f"D{QUAL_OVERALL_ROW}"

    # --- 教育達成率・バックアップ率(単純な集計) ---
    ws_health.cell(row=SIMPLE_ROW, column=1, value="教育達成率").font = BOLD
    ws_health.cell(row=SIMPLE_ROW, column=2,
                   value=f"=AVERAGE(メンバーデータ!K2:N{last_member_row})/4*100").number_format = "0.0"
    metric_cells["教育達成率"] = f"B{SIMPLE_ROW}"

    ws_health.cell(row=SIMPLE_ROW2, column=1, value="バックアップ率").font = BOLD
    ws_health.cell(row=SIMPLE_ROW2, column=2,
                   value=f'=COUNTIF(メンバーデータ!J2:J{last_member_row},"?*")/COUNTA(メンバーデータ!A2:A{last_member_row})*100').number_format = "0.0"
    metric_cells["バックアップ率"] = f"B{SIMPLE_ROW2}"

    # --- 指標別サマリー表(row6: 見出し, row7-11: 5指標, row13: 総合) ---
    table_header_row = 6
    ws_health.cell(row=table_header_row, column=1, value="指標").font = HEADER_FONT
    ws_health.cell(row=table_header_row, column=1).fill = HEADER_FILL
    ws_health.cell(row=table_header_row, column=2, value="スコア").font = HEADER_FONT
    ws_health.cell(row=table_header_row, column=2).fill = HEADER_FILL

    metric_order = ["教育達成率", "属人化耐性", "資格充足率", "バックアップ率", "変更耐性"]
    metric_start_row = table_header_row + 1  # 7
    for i, name in enumerate(metric_order):
        r = metric_start_row + i
        ws_health.cell(row=r, column=1, value=name).font = NORMAL
        ws_health.cell(row=r, column=2, value=f"={metric_cells[name]}").number_format = "0.0"

    overall_row = metric_start_row + len(metric_order) + 1  # 13
    ws_health.cell(row=overall_row, column=1, value="総合健全度スコア").font = Font(name=FONT_NAME, bold=True, size=13)
    overall_formula = "=AVERAGE(" + ",".join(f"B{metric_start_row + i}" for i in range(len(metric_order))) + ")"
    overall_cell = ws_health.cell(row=overall_row, column=2, value=overall_formula)
    overall_cell.font = Font(name=FONT_NAME, bold=True, size=13)
    overall_cell.number_format = "0.0"

    # --- 総合スコア・判定のクイック表示(row3-4、上の表と行が重ならない位置) ---
    ws_health.cell(row=3, column=1, value="総合健全度スコア").font = BOLD
    quick_cell = ws_health.cell(row=3, column=3, value=f"=B{overall_row}")
    quick_cell.font = Font(name=FONT_NAME, bold=True, size=20)
    quick_cell.number_format = "0.0"
    ws_health.cell(row=4, column=1, value="判定").font = BOLD
    ws_health.cell(row=4, column=3,
                   value='=IF(C3>=80,"安全:突発休暇があっても対応可能",IF(C3>=60,"要注意:イベント重複時は厳しい可能性","要対応:バックアップ確保が必要"))')

    # --- 今月の早番A(教育投資枠)活用回数(チーム別) ---
    investA_row = 42  # SIMPLE_ROW2(40)より後
    ws_health.cell(row=investA_row, column=1, value="今月の早番A(教育投資枠)活用回数(チーム別)").font = BOLD
    for c, h in enumerate(["チーム", "早A回数(今月)"], start=1):
        cell = ws_health.cell(row=investA_row + 1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for i, team in enumerate(["A", "B", "C", "D", "E"]):
        r = investA_row + 2 + i
        ws_health.cell(row=r, column=1, value=team)
        ws_health.cell(
            row=r, column=2,
            value=(
                f'=AVERAGEIF(\'シフト表(2026年8月)\'!$B$2:$B${last_member_row},"{team}",'
                f"'シフト表(2026年8月)'!${get_column_letter(count_col)}$2:${get_column_letter(count_col)}${last_member_row})"
            )
        )

    # --- 判定基準の説明 ---
    band_row = investA_row + 2 + 5 + 2
    ws_health.cell(row=band_row, column=1, value="判定基準").font = BOLD
    for c, h in enumerate(["スコア帯", "意味"], start=1):
        cell = ws_health.cell(row=band_row + 1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    bands = [
        ("80以上", "突発休暇が出ても、現状の体制で対応可能"),
        ("60〜79", "通常は対応可能。ただしイベント(点検・試験等)が重なる場合は要注意"),
        ("60未満", "事前にバックアップ要員の確保・出動が必要"),
    ]
    for i, (rng, meaning) in enumerate(bands):
        r = band_row + 2 + i
        ws_health.cell(row=r, column=1, value=rng)
        ws_health.cell(row=r, column=2, value=meaning)
        ws_health.column_dimensions["B"].width = 46

    # --- 教育とシフト健全度・希望休の関係についての説明 ---
    note_row2 = band_row + 2 + len(bands) + 2
    note_text = (
        "教育・資格取得を進めるほど「単独対応できる人」が増え、属人化耐性・資格充足率が上がり、"
        "シフト健全度全体が向上します。健全度が高いほど、突発的な欠勤や希望休があってもバックアップで"
        "対応でき、結果として希望休を安心して承認しやすくなります。\n"
        "今月どれだけ早番A(教育投資枠)を教育に充てたかが、3〜6ヶ月後の健全度に直結します"
        "(3ヶ月後・6ヶ月後・1年後の予測はSTEP7で追加予定)。"
    )
    note_cell = ws_health.cell(row=note_row2, column=1, value=note_text)
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    note_cell.font = Font(name=FONT_NAME, italic=True, size=10)
    ws_health.row_dimensions[note_row2].height = 60

    # ---------------- 労働時間・自動承認設定 ----------------
    ws_hours.column_dimensions["A"].width = 22
    ws_hours.column_dimensions["B"].width = 16
    ws_hours.column_dimensions["C"].width = 12
    ws_hours.column_dimensions["D"].width = 45

    ws_hours["A1"] = "労働時間チェック・自動承認設定"
    ws_hours["A1"].font = TITLE_FONT
    ws_hours.merge_cells("A1:D1")
    note = ws_hours.cell(
        row=2, column=1,
        value=(
            "週40時間を超える勤務は、シフト健全度の良し悪しとは無関係にチェックする"
            "(健全度が高くても法令上の問題は残るため)。"
        ),
    )
    note.font = Font(name=FONT_NAME, italic=True, size=9)
    ws_hours.merge_cells("A2:D2")

    ws_hours["A4"] = "自動承認設定(管理者がON/OFFできる)"
    ws_hours["A4"].font = BOLD
    settings = constraints.get("auto_approval", {})
    setting_rows = [
        ("自動承認機能", "有効" if settings.get("enabled") else "無効"),
        ("申請期限(◯日前まで)", f"{settings.get('min_days_before')}日前まで"),
        ("最低必要バックアップ人数", f"{settings.get('min_backup_headcount')}人"),
    ]
    for i, (label, value) in enumerate(setting_rows):
        ws_hours.cell(row=5 + i, column=1, value=label)
        ws_hours.cell(row=5 + i, column=2, value=value).font = BOLD

    ws_hours["A9"] = "自動承認の条件(すべて満たした場合のみ、管理者の承認なしで通る)"
    ws_hours["A9"].font = BOLD
    conditions = [
        "1. 自動承認機能が有効になっている(上記設定)",
        "2. 申請日から希望日まで、設定した日数以上ある(直前の申請は対象外)",
        "3. その週にチーム内で週40時間超えが無い(法令順守を優先し、健全度に関係なくブロック)",
        "4. 承認後もチームに単独対応できる人が十分残る(健全度への影響が小さい)",
        "5. NGペアの相手が関わる、当事者間の調整が必要な状況ではない",
    ]
    for i, c in enumerate(conditions):
        ws_hours.cell(row=10 + i, column=1, value=c)
        ws_hours.merge_cells(start_row=10 + i, start_column=1, end_row=10 + i, end_column=4)

    hour_header_row = 17
    ws_hours.cell(row=hour_header_row - 1, column=1, value="週40時間の労働時間チェック結果(2026年8月・デモ)").font = BOLD
    for c, h in enumerate(["氏名", "チーム", "週", "時間", "上限"], start=1):
        cell = ws_hours.cell(row=hour_header_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    violations = find_weekly_hour_violations(schedule_for_auto_approval, member_objs, constraints)
    r = hour_header_row + 1
    for v in violations:
        ws_hours.cell(row=r, column=1, value=v["member"])
        ws_hours.cell(row=r, column=2, value=v["team"])
        ws_hours.cell(row=r, column=3, value=v["week"])
        ws_hours.cell(row=r, column=4, value=v["hours"])
        ws_hours.cell(row=r, column=5, value=v["limit"])
        r += 1

    note2 = ws_hours.cell(
        row=r + 1, column=1,
        value=(
            "※ このデモは「早番A/早番B=8時間、夜勤=16時間」という単純な前提で計算している。"
            "実際の24時間現場では、夜勤を含むシフトは「1ヶ月単位の変形労働時間制」等で運用され、"
            "週単位ではなく月単位で平均をとるのが一般的。本チェックはあくまで週単位の簡易チェックであり、"
            "実際の適法性判断には社会保険労務士等への確認を推奨する。"
        ),
    )
    note2.font = Font(name=FONT_NAME, italic=True, size=9)
    note2.alignment = Alignment(wrap_text=True, vertical="top")
    ws_hours.merge_cells(start_row=r + 1, start_column=1, end_row=r + 3, end_column=4)
    ws_hours.row_dimensions[r + 1].height = 50

    legal_row = r + 5
    ws_hours.cell(row=legal_row, column=1, value="参考: 関連する法令メモ(簡易チェックの前提として)").font = BOLD
    for c, h in enumerate(["項目", "内容"], start=1):
        cell = ws_hours.cell(row=legal_row + 1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    legal_items = [
        ("36協定(原則)", "月45時間までの残業は毎月可能(年360時間以内)。"),
        ("36協定(特別条項)", "月45時間を「超える」残業ができるのが年6回まで。"),
        ("1日8時間超のシフト", "「変形労働時間制」を導入していれば、所定労働時間として8時間超の設定が可能。"),
        (
            "日付またぎの勤務",
            "暦日に関わらず「1つの勤務」として計算される。休憩(8時間超で1時間以上)と"
            "割増賃金を支払えば16時間拘束も可能。",
        ),
    ]
    for i, (label, text) in enumerate(legal_items):
        rr = legal_row + 2 + i
        ws_hours.cell(row=rr, column=1, value=label).font = BOLD
        cell = ws_hours.cell(row=rr, column=2, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_hours.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=4)
        ws_hours.row_dimensions[rr].height = 32
    legal_note_row = legal_row + 2 + len(legal_items) + 1
    legal_note = ws_hours.cell(
        row=legal_note_row, column=1,
        value="※ 上記は一般的な制度の概要であり、事業場ごとの労使協定・就業規則の内容が優先される。個別の適用可否は社会保険労務士等に確認すること。",
    )
    legal_note.font = Font(name=FONT_NAME, italic=True, size=9)
    legal_note.alignment = Alignment(wrap_text=True, vertical="top")
    ws_hours.merge_cells(start_row=legal_note_row, start_column=1, end_row=legal_note_row + 1, end_column=4)

    # ---------------- 頭打ち箇所と対応案(STEP7と連動) ----------------
    gap_report = instructor_gap_report(member_objs, roadmap)
    need = distinct_promotion_need(member_objs, gap_report["gaps"])

    ws_gaps.column_dimensions["A"].width = 20
    ws_gaps.column_dimensions["B"].width = 34
    ws_gaps.column_dimensions["C"].width = 34
    ws_gaps.column_dimensions["D"].width = 34
    ws_gaps.column_dimensions["E"].width = 34

    ws_gaps["A1"] = "教育担当が不在で頭打ちになっている箇所と対応案"
    ws_gaps["A1"].font = TITLE_FONT
    ws_gaps.merge_cells("A1:E1")

    ws_gaps["A3"] = "頭打ち箇所(チーム×設備の組み合わせ数)"
    ws_gaps["A3"].font = BOLD
    ws_gaps["B3"] = gap_report["gap_count"]
    ws_gaps["A4"] = "実際に昇格が必要な人数"
    ws_gaps["A4"].font = BOLD
    ws_gaps["B4"] = need["distinct_people_needed"]
    ws_gaps["C4"] = "(1人が同じチーム内の複数設備を兼任できるため、組み合わせ数より少なくて済む)"
    ws_gaps["C4"].font = Font(name=FONT_NAME, italic=True, size=9)

    summary_header_row = 6
    for c, h in enumerate(["候補者", "担当する(チーム/設備)", "件数"], start=1):
        cell = ws_gaps.cell(row=summary_header_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for i, b in enumerate(need["breakdown"]):
        r = summary_header_row + 1 + i
        covers_text = ", ".join(b["covers"])
        ws_gaps.cell(row=r, column=1, value=b["candidate"]).alignment = Alignment(vertical="top")
        cover_cell = ws_gaps.cell(row=r, column=2, value=covers_text)
        cover_cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_gaps.cell(row=r, column=3, value=len(b["covers"])).alignment = Alignment(vertical="top")
        _fit_row_height(ws_gaps, r, [(covers_text, 34)], base=15, pad=8)

    detail_start = summary_header_row + 1 + len(need["breakdown"]) + 2
    ws_gaps.cell(row=detail_start, column=1, value="箇所ごとの対応案(4案)").font = BOLD
    header_row2 = detail_start + 1
    for c, h in enumerate(["チーム/設備", "案1: 内部昇格", "案2: 他チームからの応援", "案3: 常駐化+ローテーション制", "案4: チーム編成の見直し(異動)"], start=1):
        cell = ws_gaps.cell(row=header_row2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws_gaps.row_dimensions[header_row2].height = 30

    r = header_row2 + 1
    for gap in gap_report["gaps"]:
        resolution = gap_resolution_options(member_objs, gap)
        opt_by_name = {o["案"]: o for o in resolution["options"]}
        label_cell = ws_gaps.cell(row=r, column=1, value=f"チーム{gap['team']} / {gap['equipment']}")
        label_cell.alignment = Alignment(wrap_text=True, vertical="top")
        label_cell.font = BOLD
        row_texts = [(f"チーム{gap['team']} / {gap['equipment']}", 20)]
        for col, key in zip((2, 3, 4, 5), ("内部昇格", "他チームからの応援", "常駐化+ローテーション制", "チーム編成の見直し(異動)")):
            opt = opt_by_name.get(key)
            text = opt["内容"] if opt else "(該当候補なし)"
            cell = ws_gaps.cell(row=r, column=col, value=text)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_texts.append((text, 34))
        _fit_row_height(ws_gaps, r, row_texts, base=15, pad=12)
        r += 1

    note_row = r + 1
    note_cell = ws_gaps.cell(
        row=note_row, column=1,
        value=(
            "「常駐化+ローテーション制」は、最もスキルの高いメンバーを一時的に教育専任(常日勤)にする案。"
            "ただし特定の1人に固定すると不公平感が出るため、対象者を数ヶ月ごとに交代する運用を想定している。"
        ),
    )
    note_cell.font = Font(name=FONT_NAME, italic=True, size=9)
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws_gaps.row_dimensions[note_row].height = 40

    # ---------------- 使い方(読み方6ステップ) ----------------
    ws_guide.column_dimensions["A"].width = 10
    ws_guide.column_dimensions["B"].width = 22
    ws_guide.column_dimensions["C"].width = 48
    ws_guide.column_dimensions["D"].width = 44

    ws_guide["A1"] = "このレポートの読み方(6ステップ)"
    ws_guide["A1"].font = TITLE_FONT
    ws_guide.merge_cells("A1:D1")
    ws_guide["A2"] = "スコアを見る → 原因を特定する → 対応案を選ぶ → 実行して効果を追う、という流れで使います。"
    ws_guide["A2"].font = Font(name=FONT_NAME, italic=True, size=10)
    ws_guide.merge_cells("A2:D2")

    def step_header(row, num, title):
        cell = ws_guide.cell(row=row, column=1, value=f"STEP{num}")
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell = ws_guide.cell(row=row, column=2, value=title)
        title_cell.font = Font(name=FONT_NAME, bold=True, size=12)
        title_cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws_guide.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        _fit_row_height(ws_guide, row, [(title, 24 + 46)], base=18, pad=10)

    def detail_line(row, label, text):
        label_cell = ws_guide.cell(row=row, column=2, value=label)
        label_cell.font = Font(name=FONT_NAME, bold=True, size=9, color="5A6472")
        label_cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell = ws_guide.cell(row=row, column=3, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        _fit_row_height(ws_guide, row, [(label, 22), (text, 48)])
        return cell

    r = 4
    step_header(r, 1, "総合スコアを見る")
    r += 1
    detail_line(r, "見る場所", "「シフト健全度」シート C3セル(総合健全度スコア)")
    r += 1
    detail_line(r, "判断基準", "80以上=安全 / 60〜79=要注意 / 60未満=要対応(バックアップ確保が必要)")
    r += 1
    detail_line(r, "このデモでは", "71.6 → 要注意(イベント重複時は厳しい可能性)")
    r += 2

    step_header(r, 2, "弱い指標を特定する")
    r += 1
    detail_line(r, "見る場所", "「シフト健全度」シートの指標テーブル(教育達成率・属人化耐性・資格充足率・バックアップ率・変更耐性)")
    r += 1
    ws_guide.cell(row=r, column=2, value="あるあるパターン").font = Font(name=FONT_NAME, bold=True, size=10)
    r += 1
    header_row = r
    for c, h in enumerate(["指標", "低いとき、現場ではこう見えている", "数値の目安"], start=2):
        cell = ws_guide.cell(row=header_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws_guide.row_dimensions[header_row].height = 30
    r += 1
    aruaru = [
        (
            "教育達成率が低い",
            "「新人にはまだ早い」でリーダーが仕事を抱え込み、教育の時間が取れていない",
            "70%以上: 順調 / 40〜70%: 要改善(教育投資を増やす) / 40%未満: 緊急(頭打ちの可能性大)",
        ),
        (
            "属人化耐性が低い",
            "「あの人じゃないと分からない」設備がある。その人が休むと現場が止まる",
            "80以上: 安全(3人以上対応可) / 40〜80: 要注意(1〜2人のみ) / 40未満: 危険(対応者ほぼ不在)",
        ),
        (
            "資格充足率が低い",
            "実務はこなせているが、必要な資格を持たないまま対応している人がいる(安全・法令上のリスク)",
            "80%以上: 良好 / 50〜80%: 要改善 / 50%未満: リスク大(無資格対応が多い)",
        ),
        (
            "バックアップ率が低い",
            "誰か1人が抜けると、代わりを探すのに毎回一苦労する",
            "70%以上: 良好 / 40〜70%: 要改善 / 40%未満: リスク大(誰も代われない)",
        ),
        (
            "変更耐性が低い",
            "急な欠勤・退職が出ると、その日のシフトが即座に組めなくなる",
            "80以上: 安全 / 40〜80: 要注意 / 40未満: 危険(1人抜けたら即機能不全)",
        ),
    ]
    for label, text, threshold in aruaru:
        label_cell = ws_guide.cell(row=r, column=2, value=label)
        label_cell.font = Font(name=FONT_NAME, bold=True, size=10)
        label_cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell = ws_guide.cell(row=r, column=3, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        th_cell = ws_guide.cell(row=r, column=4, value=threshold)
        th_cell.alignment = Alignment(wrap_text=True, vertical="top")
        th_cell.font = Font(name=FONT_NAME, size=9)
        _fit_row_height(ws_guide, r, [(label, 22), (text, 48), (threshold, 44)], base=15, pad=10)
        r += 1
    detail_line(r, "このデモでは", "教育達成率41.3・資格充足率50.0が低い → 上記の1番目・3番目のパターンに近い")
    r += 2

    step_header(r, 3, "原因の場所を特定する")
    r += 1
    detail_line(r, "教育達成率が低い場合", "「頭打ち箇所と対応案」シートで、どのチーム・設備に教育担当が不在かを確認する")
    r += 1
    detail_line(r, "資格充足率が低い場合", "「シフト健全度」シートの資格充足率内訳で、どの設備の資格保有者が少ないかを確認する")
    r += 1
    detail_line(r, "このデモでは", "チームD・Eに教育担当が1人もいない。電気工事士・冷凍機械責任者の保有者が少ない")
    r += 2

    step_header(r, 4, "対応案を選ぶ")
    r += 1
    detail_line(r, "見る場所", "「頭打ち箇所と対応案」シートの4案(内部昇格/他チーム応援/常駐化+ローテーション制/チーム編成の見直し)")
    r += 1
    detail_line(r, "選び方の目安", "即効性重視なら他チームからの応援、恒久的な解決を目指すなら内部昇格やチーム編成の見直し")
    r += 2

    step_header(r, 5, "個別の希望休を判断する")
    r += 1
    detail_line(r, "見る場所", "「希望休申請一覧」シートの判定列(G列)")
    r += 1
    detail_line(r, "判断の目安", "「バックアップ不要」ならそのまま承認。「要注意」「バックアップ出動が必要」なら対応を検討してから判断する")
    r += 1
    note_cell2 = detail_line(
        r, "どうしても休む場合",
        "このシステムは代替要員を自動で割り当てません。管理者がSTEP4の対応案(特に『他チームからの応援』)"
        "から即効性のある案を選び、シフト表に手動で反映してください",
    )
    note_cell2.font = Font(name=FONT_NAME, size=9, color="B04A2E")
    r += 2

    step_header(r, 6, "実行して効果を追跡する")
    r += 1
    detail_line(r, "実行", "選んだ対応策を来月のシフトに反映する")
    r += 1
    detail_line(r, "効果の確認", "python src/future_simulation.py を実行すると、3・6・12ヶ月後の健全度予測が再計算できる")
    r += 2

    step_header(r, "0", "【メンバー向け】スマホで健全度を見てから希望休を申請する")
    r += 1
    detail_line(r, "使うもの", "dashboard/ShiftDashboard.jsx(スマホのブラウザ/claude.aiのアーティファクトとして動作)")
    r += 1
    detail_line(r, "手順1", "ダッシュボードを開き、メンバービューで自分の名前を選ぶ")
    r += 1
    detail_line(r, "手順2", "画面上部に表示される『現在の健全度』を見る(このデモではメンバー全員が同じ数値を見る組織全体の健全度)")
    r += 1
    detail_line(r, "手順3", "自分のシフトを確認し、休みたい日付・理由を入力して申請する")
    r += 1
    detail_line(r, "手順4", "申請は管理者ビューに即時反映され、承認/却下の結果もこの画面に表示される")
    r += 1
    note_cell3 = detail_line(
        r, "現状の制約",
        "『この日にこの人が休んだら健全度が何点になるか』という個人単位のシミュレーションは、"
        "現バージョンでは管理者ビュー側(STEP4・5)でのみ確認できる。メンバー個人向けのシミュレーションはV2で検討",
    )
    note_cell3.font = Font(name=FONT_NAME, size=9, color="B04A2E")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"保存しました: {OUT_PATH}")


if __name__ == "__main__":
    build_workbook()
